import os
import io
import json
import time
import threading
import requests
import openpyxl
from flask import Flask, jsonify
from flask_cors import CORS
from datetime import datetime

app = Flask(__name__)
CORS(app)  # Allow all origins so your Netlify dashboard can call this

# ── CONFIG ────────────────────────────────────────────────────────────────────
# SharePoint "download" URL — converted from the share link
SHAREPOINT_URL = os.environ.get("SHAREPOINT_URL", "https://botreecoin-my.sharepoint.com/:x:/g/personal/aakash_vimalanathan_botree_co_in/IQDJDmF2hAiIQaSsId82mK8-AUdbxXLv47ZKwTaSJhOpUXQ?e=c2Yczt&download=1")
REFRESH_SECONDS = int(os.environ.get("REFRESH_SECONDS", "60"))

# ── IN-MEMORY CACHE ───────────────────────────────────────────────────────────
cache = {
    "deals": [],
    "delay_analysis": [],
    "last_updated": None,
    "status": "initializing"
}

# ── PARSE HELPERS ─────────────────────────────────────────────────────────────
def safe_str(v):
    if v is None: return ""
    s = str(v).strip()
    return "" if s in ("None", "nan", "\xa0") else s

def safe_float(v):
    try: return round(float(v), 6)
    except: return 0.0

def safe_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.strftime("%Y-%m-%d")
    s = safe_str(v)
    return s[:10] if s else None

def normalize_status(s):
    s = safe_str(s)
    if s in ("Scaled Up", "Scaled up"): return "Scaled Up"
    if s == "Drop":                      return "Dropped"
    if s in ("Dev", "Dev/Pilot"):        return "Dev/Pilot"
    if s == "No Response as of 12 June": return "Dropped"
    if not s:                            return "Not Started"
    return s

# ── PARSE MAIN SHEET ─────────────────────────────────────────────────────────
def parse_main_sheet(ws):
    # Find header row (contains "Customer Name")
    hdr_row = None
    for row in ws.iter_rows():
        vals = [safe_str(c.value) for c in row]
        if "Customer Name" in vals:
            hdr_row = row
            break
    if not hdr_row:
        return []

    headers = [safe_str(c.value) for c in hdr_row]
    hdr_idx = hdr_row[0].row

    def col(name):
        for i, h in enumerate(headers):
            if name in h:
                return i
        return -1

    ci = {
        "sno": col("S.No"), "customer": col("Customer Name"),
        "salesTeam": col("Sales Team"), "ch": col("CH"),
        "product": col("Product"), "orderType": col("Order Type"),
        "orderValue": col("Order Value\n(Rs Crs)"),
        "arrValue": col("Order Value\n-ARR"), "oneTimeValue": col("One Time\n(Rs Crs)"),
        "arrInvoiced": col("ARR Invoiced"), "arrBalance": col("ARR Balance"),
        "oneTimeInvoiced": col("One Time  Invoiced"),
        "dateVerbal": col("Verbal Signoff"), "datePO": col("PO"),
        "dateContract": col("Contract"), "dateDiscovery": col("Scoping"),
        "dateMasterData": col("Master Data"), "dateDev": col("Development Start"),
        "dateUAT": col("UAT"), "dateGoLive": col("Go Live"),
        "dateScaling": col("Scaling"), "sdhComment": col("SDH Comment"),
        "reviewComment": col("Review Comments"), "status": col("Status"),
        "delayReason": col("Reason for delay"),
    }

    deals = []
    for row in ws.iter_rows(min_row=hdr_idx + 2, values_only=True):
        def g(key): return row[ci[key]] if ci.get(key, -1) >= 0 and ci[key] < len(row) else None
        sno = safe_str(g("sno"))
        if not sno or not sno.isdigit():
            continue
        deals.append({
            "sno": int(sno),
            "customer": safe_str(g("customer")) or "—",
            "salesTeam": safe_str(g("salesTeam")),
            "ch": safe_str(g("ch")),
            "product": safe_str(g("product")) or "—",
            "orderType": safe_str(g("orderType")),
            "orderValue": safe_float(g("orderValue")),
            "arrValue": safe_float(g("arrValue")),
            "oneTimeValue": safe_float(g("oneTimeValue")),
            "arrInvoiced": safe_float(g("arrInvoiced")),
            "arrBalance": safe_float(g("arrBalance")),
            "oneTimeInvoiced": safe_float(g("oneTimeInvoiced")),
            "dateVerbal": safe_date(g("dateVerbal")),
            "datePO": safe_date(g("datePO")),
            "dateContract": safe_date(g("dateContract")),
            "dateDiscovery": safe_date(g("dateDiscovery")),
            "dateMasterData": safe_date(g("dateMasterData")),
            "dateDev": safe_date(g("dateDev")),
            "dateUAT": safe_date(g("dateUAT")),
            "dateGoLive": safe_date(g("dateGoLive")),
            "dateScaling": safe_date(g("dateScaling")),
            "sdhComment": safe_str(g("sdhComment")),
            "reviewComment": safe_str(g("reviewComment")),
            "status": normalize_status(g("status")),
            "delayReason": safe_str(g("delayReason")),
        })
    return deals

# ── PARSE DELAY ANALYSIS SHEET ────────────────────────────────────────────────
def parse_delay_sheet(ws):
    rows = []
    started = False
    for row in ws.iter_rows(values_only=True):
        vals = [safe_str(v) for v in row]
        if "Reason for Delay" in vals:
            started = True
            continue
        if not started:
            continue
        reason = vals[0] if vals else ""
        if not reason or reason in ("TOTAL", "▶  MONTHLY REVENUE LOSS", "CUSTOMER BREAKDOWN BY DELAY REASON"):
            continue
        try:
            rows.append({
                "reason": reason,
                "customers": int(float(vals[1])) if vals[1] else 0,
                "totalARR": safe_float(vals[2]),
                "monthlyLoss": safe_float(vals[3]),
                "pctTotal": safe_float(vals[4].replace("%","")) if "%" in (vals[4] or "") else safe_float(vals[4]),
            })
        except:
            continue
    return rows

# ── FETCH & REFRESH ───────────────────────────────────────────────────────────
def fetch_and_parse():
    global cache
    try:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Fetching Excel from SharePoint…")
        headers = {
            "User-Agent": "Mozilla/5.0",
            "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
        }
        resp = requests.get(SHAREPOINT_URL, headers=headers, timeout=30, allow_redirects=True)
        resp.raise_for_status()

        wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)

        # Find main sheet
        main_ws = None
        for name in wb.sheetnames:
            if "New Orders" in name and "(2)" not in name:
                main_ws = wb[name]
                break
        if not main_ws:
            main_ws = wb.active

        deals = parse_main_sheet(main_ws)

        # Delay analysis sheet
        delay = []
        if "Delay Analysis" in wb.sheetnames:
            delay = parse_delay_sheet(wb["Delay Analysis"])

        cache["deals"] = deals
        cache["delay_analysis"] = delay
        cache["last_updated"] = datetime.now().isoformat()
        cache["status"] = "ok"
        print(f"[{datetime.now().strftime('%H:%M:%S')}] ✓ Loaded {len(deals)} deals")

    except Exception as e:
        cache["status"] = f"error: {str(e)}"
        print(f"[{datetime.now().strftime('%H:%M:%S')}] ✗ Error: {e}")

def background_refresh():
    while True:
        fetch_and_parse()
        time.sleep(REFRESH_SECONDS)

# ── ROUTES ────────────────────────────────────────────────────────────────────
@app.route("/")
def health():
    return jsonify({
        "status": cache["status"],
        "last_updated": cache["last_updated"],
        "deals_count": len(cache["deals"]),
        "refresh_seconds": REFRESH_SECONDS,
    })

@app.route("/api/data")
def get_data():
    return jsonify({
        "deals": cache["deals"],
        "delay_analysis": cache["delay_analysis"],
        "last_updated": cache["last_updated"],
        "status": cache["status"],
        "deals_count": len(cache["deals"]),
    })

@app.route("/api/refresh", methods=["POST"])
def manual_refresh():
    thread = threading.Thread(target=fetch_and_parse)
    thread.daemon = True
    thread.start()
    return jsonify({"message": "Refresh triggered"})

# ── STARTUP ───────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    # Initial fetch
    fetch_and_parse()
    # Background refresh thread
    t = threading.Thread(target=background_refresh)
    t.daemon = True
    t.start()
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
else:
    # For gunicorn (Render uses this)
    fetch_and_parse()
    t = threading.Thread(target=background_refresh)
    t.daemon = True
    t.start()
